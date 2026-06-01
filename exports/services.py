"""Exportação de lotes para Excel (openpyxl)."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from audit.models import Log
from audit.services import log_action

CABECALHOS = ["#", "ICCID", "Tentativas", "Status leitura", "Corrigido", "Data leitura"]


def gerar_workbook(lote):
    """Monta o Workbook com os chips ativos do lote."""
    wb = Workbook()
    ws = wb.active
    ws.title = "ICCIDs"

    ws.append([f"Lote: {lote.nome}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    ws.append(CABECALHOS)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for chip in lote.chips.filter(is_active=True).order_by("sequencia"):
        ws.append([
            chip.sequencia,
            chip.iccid,
            chip.tentativas,
            chip.get_status_leitura_display(),
            "Sim" if chip.corrigido_manualmente else "Não",
            chip.data_leitura.strftime("%d/%m/%Y %H:%M"),
        ])

    # Largura aproximada das colunas.
    larguras = [6, 26, 12, 16, 12, 20]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = largura

    return wb


def exportar_lote_excel(lote, usuario):
    """Gera os bytes do arquivo .xlsx e registra a auditoria."""
    wb = gerar_workbook(lote)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    log_action(
        Log.Acao.EXPORTAR,
        usuario=usuario,
        resultado={"lote_id": str(lote.id), "quantidade": lote.quantidade},
    )
    return buffer.getvalue()
